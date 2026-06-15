"""
build_human_review_sampling.py

Constructs a three-layer human-review sample from the Document Status Index
and allocates document-level review packets to reviewers.

This version creates complete Spanish-language reviewer workspaces:

outputs/human_review/reviewer_packets/
    reviewer_1_packet/
        reviewer_1_packet.xlsx
        <document_id>/
            <original HTR filename>
            <original GT filename>
    ...

The Excel reviewer sheet is fully in Spanish and includes:
    - ID de revisión
    - ID del documento
    - Estilo
    - one Spanish review-domain column per review dimension
    - Estado de revisión

Dropdown options are Spanish "Etiqueta - definición" strings. Conditional
formatting applies traffic-light colouring based on the selected option.

This script is designed to run after the diagnostics/report stage has produced
a document status or operational metrics CSV containing, ideally:
    doc_id, style, risk_band, document_status/governance_status,
    htr_path, gt_path

The preferred input is:
    outputs/ml_outputs/document_status_index.csv

The script is reproducible when the input CSV, sampling parameters, and random
seed remain unchanged.
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import random
import re
import shutil
from collections import defaultdict
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.utils import get_column_letter
    from openpyxl.workbook.defined_name import DefinedName
except ImportError as exc:
    raise SystemExit(
        "This script requires openpyxl. Install it with `pip install openpyxl` "
        "or run it in the project environment where openpyxl is available."
    ) from exc


# ---------------------------------------------------------------------
# User-controlled parameters
# ---------------------------------------------------------------------

OUTPUT_DIR = Path("outputs") / "human_review"
METRICS_CSV: Path | None = None

NUMBER_OF_REVIEWERS = 5
GOVERNANCE_VALIDATION_TOTAL = 75
RANDOM_CONTROL_DOCS_PER_STYLE = 3
DIAGNOSTIC_CASE_STUDY_TOTAL = 10
RANDOM_SEED = 42

# "even" validates the whole cell distribution; "random" samples randomly;
# "extreme" prioritises most severe documents within each cell.
GOVERNANCE_SELECTION_METHOD = "even"

# If htr_path/gt_path are absent from the input CSV, these roots are used as
# fallbacks. Normally the explicit htr_path and gt_path columns should be used.
RAW_DATA_ROOT = Path("data") / "raw"
GROUND_TRUTH_ROOT = RAW_DATA_ROOT / "ground_truths" / "Corpus_GT"


# ---------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------

DEFAULT_METRICS_CANDIDATES = [
    Path("outputs/ml_outputs/document_status_index.csv"),
    Path("logs/posthoc/corpus_report_tables/document_status_index.csv"),
    Path("logs/posthoc/corpus_report_tables/operational_document_metrics.csv"),
    Path("logs/posthoc/corpus_report_tables/rag_document_quality.csv"),
]

RISK_BANDS = ["Low", "Medium", "High"]

FILL_GREEN = "D9EAD3"
FILL_AMBER = "FFF2CC"
FILL_RED = "F4CCCC"
FILL_BLUE = "D9EAF7"
FILL_GREY = "E7E6E6"
FILL_HEADER = "1F4E79"
FILL_PURPLE = "EADCF8"
FILL_DESCRIPTION = "F3F6FA"
FILL_READONLY = "EDEDED"

SYMBOL_GREEN = "🟢"
SYMBOL_AMBER = "🟡"
SYMBOL_RED = "🔴"
SYMBOL_GREY = "⚪"


REVIEW_DIMENSIONS = [
    "fiabilidad_general_transcripcion",
    "correspondencia_con_gt",
    "severidad_omisiones",
    "conservacion_disposicion_documento",
    "utilidad_academica",
    "confianza_revisor",
    "estado_revision",
]

SPANISH_COLUMNS = [
    "ID de revisión",
    "ID del documento",
    "Estilo",
    "Fiabilidad General de la Transcripción",
    "Continuidad de la Correspondencia con el GT",
    "Severidad de las Omisiones",
    "Conservación de la Estructura Documental",
    "Utilidad Académica",
    "Confianza del Revisor",
    "Estado de Revisión",
]

REVIEW_COLUMNS_INTERNAL = [
    "review_id",
    "doc_id",
    "style",
    *REVIEW_DIMENSIONS,
]

DIMENSION_SPANISH_HEADERS = {
    "fiabilidad_general_transcripcion": "Fiabilidad General de la Transcripción",
    "correspondencia_con_gt": "Continuidad de la Correspondencia con el GT",
    "severidad_omisiones": "Severidad de las Omisiones",
    "conservacion_disposicion_documento": "Conservación de la Estructura Documental",
    "utilidad_academica": "Utilidad Académica",
    "confianza_revisor": "Confianza del Revisor",
    "estado_revision": "Estado de Revisión",
}

DIMENSION_ENGLISH_HEADERS = {
    "fiabilidad_general_transcripcion": "Overall Transcription Reliability",
    "correspondencia_con_gt": "Continuity of Correspondence with the GT",
    "severidad_omisiones": "Omission Severity",
    "conservacion_disposicion_documento": "Preservation of Document Structure",
    "utilidad_academica": "Scholarly Usability",
    "confianza_revisor": "Reviewer Confidence",
    "estado_revision": "Review Status",
}

DIMENSION_DESCRIPTIONS = {
    "fiabilidad_general_transcripcion": (
        "Definición: Esta categoría evalúa hasta qué punto la transcripción HTR representa fielmente el contenido del GT. "
        "Se centra en la fiabilidad general del texto y en si puede considerarse una representación adecuada del documento original.\n"
        "Pregunta: ¿Hasta qué punto la transcripción HTR representa fielmente el GT?"
    ),
    "correspondencia_con_gt": (
        "Definición: Esta categoría evalúa si es posible relacionar la transcripción HTR con el GT de forma continua a lo largo del documento. "
        "Se centra en la continuidad de la correspondencia textual entre ambos textos y no en la estructura del documento ni en la disposición de las líneas.\n"
        "Ejemplo: Puede ocurrir que la omisión o inserción de caracteres o palabras provoque que la correspondencia entre HTR y GT se interrumpa repetidamente, "
        "generando numerosos tramos cortos de correspondencia en lugar de tramos largos y continuos.\n"
        "Pregunta: ¿La correspondencia textual entre la transcripción HTR y el GT se mantiene de forma continua a lo largo del documento?"
    ),
    "severidad_omisiones": (
        "Definición: Esta categoría evalúa la frecuencia y el impacto de palabras, frases o líneas que están presentes en el GT "
        "pero ausentes en la transcripción HTR.\n"
        "Pregunta: ¿Con qué frecuencia se omiten palabras, frases o líneas?"
    ),
    "conservacion_disposicion_documento": (
        "Definición: Esta categoría evalúa si la organización estructural del documento se conserva en la transcripción HTR. "
        "Se refiere a aspectos como líneas omitidas, líneas duplicadas, líneas fusionadas o divididas incorrectamente y alteraciones del orden de lectura. "
        "No evalúa la continuidad de la correspondencia textual entre HTR y GT.\n"
        "Ejemplo: Si dos líneas independientes del GT aparecen fusionadas en una única línea en la transcripción HTR, existe un problema estructural aunque el contenido textual siga siendo reconocible.\n"
        "Pregunta: ¿La estructura y organización general del documento se conservan correctamente en la transcripción HTR?"
    ),
    "utilidad_academica": (
        "Definición: Esta categoría evalúa si la transcripción HTR puede utilizarse para investigación histórica sin necesidad de realizar "
        "correcciones importantes. Se centra en la utilidad práctica de la transcripción para el trabajo académico.\n"
        "Pregunta: ¿Podría utilizar esta transcripción para investigación histórica sin realizar correcciones importantes?"
    ),
    "confianza_revisor": (
        "Definición: Esta categoría evalúa el grado de confianza que tiene el revisor en su propia valoración. Una confianza baja puede "
        "reflejar dificultades de lectura, ambigüedad paleográfica o incertidumbre interpretativa.\n"
        "Pregunta: ¿Qué grado de confianza tiene en esta evaluación?"
    ),
    "estado_revision": (
        "Definición: Esta categoría indica el estado actual del proceso de revisión del documento.\n"
        "Pregunta: ¿Cuál es el estado actual de la revisión?"
    ),
}

DIMENSION_DESCRIPTIONS_EN = {
    "fiabilidad_general_transcripcion": (
        "Definition: This category assesses the extent to which the HTR transcription faithfully represents the content of the GT. "
        "It focuses on the general reliability of the text and whether it can be considered an adequate representation of the original document.\n"
        "Question: To what extent does the HTR transcription faithfully represent the GT?"
    ),
    "correspondencia_con_gt": (
        "Definition: This category assesses whether the HTR transcription can be related to the GT continuously throughout the document. "
        "It focuses on the continuity of textual correspondence between the two texts, not on document structure or line layout.\n"
        "Example: An omission or insertion of characters or words may cause the correspondence between HTR and GT to be interrupted repeatedly, "
        "creating many short correspondence spans rather than long continuous spans.\n"
        "Question: Is the textual correspondence between the HTR transcription and the GT maintained continuously throughout the document?"
    ),
    "severidad_omisiones": (
        "Definition: This category assesses the frequency and impact of words, phrases, or lines that are present in the GT but absent from the HTR transcription.\n"
        "Question: How frequently are words, phrases, or lines omitted?"
    ),
    "conservacion_disposicion_documento": (
        "Definition: This category assesses whether the structural organisation of the document is preserved in the HTR transcription. "
        "It refers to issues such as omitted lines, duplicated lines, incorrectly merged or split lines, and changes to reading order. "
        "It does not assess the continuity of textual correspondence between HTR and GT.\n"
        "Example: If two independent GT lines appear merged into a single line in the HTR transcription, this is a structural problem even if much of the textual content remains recognisable.\n"
        "Question: Is the general structure and organisation of the document preserved correctly in the HTR transcription?"
    ),
    "utilidad_academica": (
        "Definition: This category assesses whether the HTR transcription can be used for historical research without major correction. "
        "It focuses on the practical usefulness of the transcription for scholarly work.\n"
        "Question: Could this transcription be used for historical research without major correction?"
    ),
    "confianza_revisor": (
        "Definition: This category records the reviewer's confidence in their own assessment. Low confidence may reflect reading difficulty, palaeographic ambiguity, or interpretive uncertainty.\n"
        "Question: How confident are you in this assessment?"
    ),
    "estado_revision": (
        "Definition: This category records the current status of the document review process.\n"
        "Question: What is the current review status?"
    ),
}

LABEL_SETS: dict[str, list[dict[str, str]]] = {
    "fiabilidad_general_transcripcion": [
        {
            "label_es": "Fiable",
            "meaning_es": "La transcripción HTR es, en términos generales, una representación fiable del GT.",
            "label_en": "Reliable",
            "meaning_en": "The HTR is broadly trustworthy as a transcription of the GT.",
            "colour": FILL_GREEN,
        },
        {
            "label_es": "Mayormente fiable",
            "meaning_es": "La transcripción HTR es utilizable, aunque contiene errores apreciables.",
            "label_en": "Mostly reliable",
            "meaning_en": "The HTR is usable, although it contains noticeable errors.",
            "colour": FILL_AMBER,
        },
        {
            "label_es": "Presenta problemas significativos",
            "meaning_es": "La transcripción HTR contiene errores importantes que afectan su utilización.",
            "label_en": "Has significant problems",
            "meaning_en": "The HTR contains important errors that affect its use.",
            "colour": FILL_RED,
        },
        {
            "label_es": "Gravemente errónea",
            "meaning_es": "La transcripción HTR resulta frecuentemente inutilizable o claramente errónea.",
            "label_en": "Seriously erroneous",
            "meaning_en": "The HTR is frequently unusable or clearly erroneous.",
            "colour": FILL_RED,
        },
    ],
    "correspondencia_con_gt": [
        {
            "label_es": "Correspondencia continua mantenida",
            "meaning_es": "La correspondencia entre HTR y GT se mantiene de forma continua durante casi todo el documento.",
            "label_en": "Continuous correspondence maintained",
            "meaning_en": "The correspondence between HTR and GT remains continuous through almost the entire document.",
            "colour": FILL_GREEN,
        },
        {
            "label_es": "Interrupciones locales de continuidad",
            "meaning_es": "Existen algunas interrupciones de correspondencia, pero predominan los tramos largos y continuos.",
            "label_en": "Local continuity interruptions",
            "meaning_en": "There are some interruptions in correspondence, but long continuous spans still predominate.",
            "colour": FILL_AMBER,
        },
        {
            "label_es": "Interrupciones frecuentes de continuidad",
            "meaning_es": "La correspondencia se rompe repetidamente y aparecen numerosos tramos cortos de correspondencia.",
            "label_en": "Frequent continuity interruptions",
            "meaning_en": "The correspondence breaks repeatedly, producing many short correspondence spans.",
            "colour": FILL_RED,
        },
        {
            "label_es": "Pérdida extensa de continuidad",
            "meaning_es": "Gran parte del documento carece de correspondencia continua entre HTR y GT.",
            "label_en": "Extensive loss of continuity",
            "meaning_en": "Large portions of the document lack continuous correspondence between HTR and GT.",
            "colour": FILL_RED,
        },
    ],
    "severidad_omisiones": [
        {
            "label_es": "Omisiones mínimas",
            "meaning_es": "Las omisiones son escasas y tienen poco impacto sobre la comprensión.",
            "label_en": "Minimal omissions",
            "meaning_en": "Omissions are scarce and have little impact on comprehension.",
            "colour": FILL_GREEN,
        },
        {
            "label_es": "Omisiones apreciables",
            "meaning_es": "Existen omisiones visibles que afectan parcialmente la lectura.",
            "label_en": "Noticeable omissions",
            "meaning_en": "Visible omissions partially affect reading.",
            "colour": FILL_AMBER,
        },
        {
            "label_es": "Omisiones frecuentes",
            "meaning_es": "Las omisiones aparecen con regularidad y dificultan la interpretación.",
            "label_en": "Frequent omissions",
            "meaning_en": "Omissions appear regularly and hinder interpretation.",
            "colour": FILL_RED,
        },
        {
            "label_es": "Omisiones severas",
            "meaning_es": "Las omisiones son numerosas y comprometen gravemente la utilidad de la transcripción.",
            "label_en": "Severe omissions",
            "meaning_en": "Omissions are numerous and seriously compromise the usefulness of the transcription.",
            "colour": FILL_RED,
        },
    ],
    "conservacion_disposicion_documento": [
        {
            "label_es": "Estructura documental conservada",
            "meaning_es": "La organización general del documento se conserva correctamente.",
            "label_en": "Document structure preserved",
            "meaning_en": "The overall organisation of the document is preserved correctly.",
            "colour": FILL_GREEN,
        },
        {
            "label_es": "Problemas estructurales menores",
            "meaning_es": "Existen algunos problemas aislados en la organización estructural del documento.",
            "label_en": "Minor structural problems",
            "meaning_en": "There are some isolated problems in the structural organisation of the document.",
            "colour": FILL_AMBER,
        },
        {
            "label_es": "Problemas estructurales recurrentes",
            "meaning_es": "Los problemas estructurales aparecen repetidamente a lo largo del documento.",
            "label_en": "Recurrent structural problems",
            "meaning_en": "Structural problems appear repeatedly throughout the document.",
            "colour": FILL_RED,
        },
        {
            "label_es": "Estructura documental gravemente alterada",
            "meaning_es": "La organización general del documento está ampliamente comprometida.",
            "label_en": "Document structure seriously altered",
            "meaning_en": "The overall organisation of the document is severely compromised.",
            "colour": FILL_RED,
        },
    ],
    "utilidad_academica": [
        {
            "label_es": "Plenamente utilizable",
            "meaning_es": "Puede utilizarse directamente para investigación histórica.",
            "label_en": "Fully usable",
            "meaning_en": "It can be used directly for historical research.",
            "colour": FILL_GREEN,
        },
        {
            "label_es": "Utilizable con precaución",
            "meaning_es": "Puede utilizarse, aunque requiere atención a determinadas secciones.",
            "label_en": "Usable with caution",
            "meaning_en": "It can be used, although some sections require caution.",
            "colour": FILL_AMBER,
        },
        {
            "label_es": "Requiere correcciones sustanciales",
            "meaning_es": "Es necesario corregir partes importantes antes de utilizarla.",
            "label_en": "Requires substantial correction",
            "meaning_en": "Important parts need correction before use.",
            "colour": FILL_RED,
        },
        {
            "label_es": "No es fiable para su uso",
            "meaning_es": "No puede utilizarse con seguridad sin una revisión extensa.",
            "label_en": "Not reliable for use",
            "meaning_en": "It cannot be used safely without extensive review.",
            "colour": FILL_RED,
        },
    ],
    "confianza_revisor": [
        {
            "label_es": "Alta",
            "meaning_es": "Confío plenamente en mi evaluación.",
            "label_en": "High",
            "meaning_en": "I am fully confident in my assessment.",
            "colour": FILL_GREEN,
        },
        {
            "label_es": "Media",
            "meaning_es": "Tengo algunas dudas, pero considero que mi evaluación es razonablemente fiable.",
            "label_en": "Medium",
            "meaning_en": "I have some doubts, but consider my assessment reasonably reliable.",
            "colour": FILL_AMBER,
        },
        {
            "label_es": "Baja",
            "meaning_es": "Mi evaluación es incierta debido a dificultades de lectura o interpretación.",
            "label_en": "Low",
            "meaning_en": "My assessment is uncertain due to reading or interpretive difficulties.",
            "colour": FILL_RED,
        },
    ],
    "estado_revision": [
        {
            "label_es": "Sin revisar",
            "meaning_es": "La revisión aún no se ha realizado.",
            "label_en": "Not reviewed",
            "meaning_en": "The review has not yet been carried out.",
            "colour": FILL_GREY,
        },
        {
            "label_es": "En revisión",
            "meaning_es": "La revisión está en curso.",
            "label_en": "In progress",
            "meaning_en": "The review is in progress.",
            "colour": FILL_AMBER,
        },
        {
            "label_es": "Revisión completada",
            "meaning_es": "La revisión se ha completado.",
            "label_en": "Completed",
            "meaning_en": "The review has been completed.",
            "colour": FILL_GREEN,
        },
    ],
}

MASTER_COLUMNS = [
    "review_id", "reviewer_id", "doc_id", "packet_doc_id", "style",
    "selection_layer", "selection_reason", "review_order",
    "governance_status", "risk_band", "risk_score", "risk_rank_within_style",
    "escalation_flags", "cer_norm", "wer_norm", "indel_disruption_rate",
    "continuity", "adjusted_line_delta_abs", "line_structure_band",
    "line_structure_flag", "diagnostic_case_score",
    "htr_path", "gt_path", "image_path_or_url", "gt_path_or_url", "htr_path_or_url",
]


# ---------------------------------------------------------------------
# CSV helpers
# ---------------------------------------------------------------------

def read_csv_dicts(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def write_csv_dicts(path: Path, rows: list[dict[str, Any]], columns: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def as_float(row: dict[str, Any], key: str, default: float = 0.0) -> float:
    try:
        val = row.get(key, default)
        if val is None or val == "":
            return default
        x = float(str(val).replace(",", ""))
        if math.isnan(x) or math.isinf(x):
            return default
        return x
    except Exception:
        return default


def norm_band(value: str) -> str:
    v = str(value or "").strip().lower()
    if v == "low":
        return "Low"
    if v == "medium":
        return "Medium"
    if v == "high":
        return "High"
    return "Unassigned"


def norm_governance_status(value: str, fallback_risk_band: str = "Unassigned") -> str:
    v = str(value or "").strip().lower()
    if v == "stable":
        return "Stable"
    if v == "review":
        return "Review"
    if v == "exclude":
        return "Exclude"
    if fallback_risk_band == "Low":
        return "Stable"
    if fallback_risk_band == "Medium":
        return "Review"
    if fallback_risk_band == "High":
        return "Exclude"
    return "Unassigned"


def _first_non_empty(row: dict[str, Any], *keys: str) -> str:
    for key in keys:
        val = row.get(key, "")
        if str(val).strip():
            return str(val).strip()
    return ""


def parse_compound_doc_id(doc_id: str) -> tuple[str, str]:
    """Return packet folder id and HTR filename from a compound doc_id.

    Example:
        AGI_CONTRATACION_1170A_N10_160:AGI..._HTR.txt
    becomes:
        packet_doc_id = AGI_CONTRATACION_1170A_N10_160
        htr_filename = AGI..._HTR.txt
    """
    doc_id = str(doc_id or "").strip()
    if ":" in doc_id:
        left, right = doc_id.split(":", 1)
        return left.strip(), Path(right.strip()).name
    return doc_id, Path(doc_id).name


def expected_gt_filename(packet_doc_id: str) -> str:
    return f"{packet_doc_id}_GT.txt"


def normalise_doc_rows(rows: list[dict[str, str]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    aliases = {
        "CER": "cer_norm",
        "WER": "wer_norm",
        "IDR": "indel_disruption_rate",
        "Continuity": "continuity",
        "Adjusted line delta": "adjusted_line_delta_abs",
        "Risk band": "risk_band",
        "Risk Band": "risk_band",
        "Index-derived risk band": "risk_band",
        "Risk score": "risk_score",
        "Risk Score": "risk_score",
        "Risk rank within style": "risk_rank_within_style",
        "Line-structure band": "line_structure_band",
        "Line-structure flag": "line_structure_flag",
        "Governance status": "governance_status",
        "Document status": "document_status",
        "Status": "document_status",
        "Escalation flags": "escalation_flags",
        "HTR path": "htr_path",
        "GT path": "gt_path",
        "htr_path_or_url": "htr_path",
        "gt_path_or_url": "gt_path",
        "HTR": "htr_path",
        "GT": "gt_path",
    }
    for r in rows:
        row: dict[str, Any] = dict(r)
        for source, target in aliases.items():
            if source in row and not str(row.get(target, "")).strip():
                row[target] = row.get(source, "")

        row["risk_band"] = norm_band(_first_non_empty(row, "risk_band", "Risk band", "Risk Band"))
        status_raw = _first_non_empty(row, "document_status", "governance_status", "Document status", "Governance status", "Status")
        row["governance_status"] = norm_governance_status(status_raw, row["risk_band"])
        row["document_status"] = row["governance_status"]
        row["style"] = str(row.get("style") or row.get("Style") or "UNKNOWN").strip() or "UNKNOWN"
        row["doc_id"] = _first_non_empty(row, "doc_id", "Doc ID", "Document ID", "id", "filename")
        row["packet_doc_id"], row["expected_htr_filename"] = parse_compound_doc_id(row["doc_id"])
        row["expected_gt_filename"] = expected_gt_filename(row["packet_doc_id"])

        if row["doc_id"]:
            out.append(row)
    return out


# ---------------------------------------------------------------------
# Sampling logic
# ---------------------------------------------------------------------

def evenly_spaced_sample(rows: list[dict[str, Any]], n: int, sort_key: str = "risk_score") -> list[dict[str, Any]]:
    if n <= 0 or not rows:
        return []
    ordered = sorted(rows, key=lambda r: as_float(r, sort_key), reverse=True)
    if len(ordered) <= n:
        return list(ordered)
    if n == 1:
        return [ordered[len(ordered) // 2]]
    idxs = [round(i * (len(ordered) - 1) / (n - 1)) for i in range(n)]
    picked: list[dict[str, Any]] = []
    seen: set[int] = set()
    for idx in idxs:
        if idx not in seen:
            picked.append(ordered[idx])
            seen.add(idx)
    j = 0
    while len(picked) < n and j < len(ordered):
        if j not in seen:
            picked.append(ordered[j])
            seen.add(j)
        j += 1
    return picked[:n]


def governance_validation_sample(rows: list[dict[str, Any]], target_total: int, method: str, rng: random.Random) -> list[dict[str, Any]]:
    by_cell: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for r in rows:
        status = r.get("governance_status", "Unassigned")
        band = r.get("risk_band", "Unassigned")
        if status == "Unassigned" or band == "Unassigned":
            continue
        by_cell[(r["style"], status, band)].append(r)

    cell_pools: dict[tuple[str, str, str], list[dict[str, Any]]] = {}
    for cell, cell_rows in by_cell.items():
        if method == "random":
            pool = list(cell_rows)
            rng.shuffle(pool)
        elif method == "extreme":
            pool = sorted(cell_rows, key=lambda r: as_float(r, "risk_score"), reverse=True)
        else:
            pool = evenly_spaced_sample(cell_rows, len(cell_rows), "risk_score")
        cell_pools[cell] = pool

    selected: list[dict[str, Any]] = []
    selected_ids: set[str] = set()
    cells = sorted(cell_pools.keys(), key=lambda c: (c[0], c[1], c[2]))
    pointers = {cell: 0 for cell in cells}

    while len(selected) < target_total:
        progressed = False
        for cell in cells:
            if len(selected) >= target_total:
                break
            pool = cell_pools[cell]
            idx = pointers[cell]
            while idx < len(pool) and pool[idx]["doc_id"] in selected_ids:
                idx += 1
            pointers[cell] = idx
            if idx >= len(pool):
                continue
            style, status, band = cell
            q = dict(pool[idx])
            pointers[cell] += 1
            selected_ids.add(q["doc_id"])
            q["selection_layer"] = "governance_validation"
            q["selection_reason"] = (
                f"style={style}; governance_status={status}; risk_band={band}; "
                f"method={method}; target_total={target_total}"
            )
            selected.append(q)
            progressed = True
        if not progressed:
            break
    return selected


def random_control_sample(rows: list[dict[str, Any]], controls_per_style: int, excluded_doc_ids: set[str], rng: random.Random) -> list[dict[str, Any]]:
    selected: list[dict[str, Any]] = []
    by_style: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for r in rows:
        if r["doc_id"] not in excluded_doc_ids:
            by_style[r["style"]].append(r)
    for style in sorted(by_style):
        pool = by_style[style]
        picks = rng.sample(pool, min(controls_per_style, len(pool)))
        for p in picks:
            q = dict(p)
            q["selection_layer"] = "random_control"
            q["selection_reason"] = f"style_stratified_random_control; controls_per_style={controls_per_style}"
            selected.append(q)
    return selected


def diagnostic_case_score(row: dict[str, Any]) -> tuple[float, list[str]]:
    reasons: list[str] = []
    score = 0.0

    cer = as_float(row, "cer_norm")
    idr = as_float(row, "indel_disruption_rate")
    continuity = as_float(row, "continuity", default=1.0)
    line_delta = as_float(row, "adjusted_line_delta_abs")
    anomaly = as_float(row, "basic_anomaly_density")
    orth = as_float(row, "orthographic_violation_density")

    if str(row.get("line_structure_flag", "")).lower() == "review_line_structure":
        score += 5
        reasons.append("review_line_structure")
    if str(row.get("line_structure_band", "")).lower() == "severe instability":
        score += 5
        reasons.append("severe_line_instability")
    if line_delta >= 4:
        score += min(line_delta, 10)
        reasons.append("high_adjusted_line_delta")
    if continuity <= 0.15:
        score += 4
        reasons.append("very_low_continuity")
    elif continuity <= 0.35:
        score += 2
        reasons.append("low_continuity")
    if idr >= 0.15:
        score += 3
        reasons.append("high_idr")
    elif idr >= 0.08:
        score += 1.5
        reasons.append("elevated_idr")
    if cer >= 0.35:
        score += 3
        reasons.append("high_cer")
    if anomaly > 0:
        score += min(anomaly / 5.0, 2)
        reasons.append("surface_anomaly_density")
    if orth >= 10:
        score += min(orth / 10.0, 2)
        reasons.append("orthographic_violation_density")

    return score, reasons


def diagnostic_case_studies(rows: list[dict[str, Any]], total: int, excluded_doc_ids: set[str]) -> list[dict[str, Any]]:
    candidates: list[tuple[float, list[str], dict[str, Any]]] = []
    for r in rows:
        if r["doc_id"] in excluded_doc_ids:
            continue
        score, reasons = diagnostic_case_score(r)
        if score > 0:
            candidates.append((score, reasons, r))
    candidates.sort(key=lambda x: x[0], reverse=True)

    selected: list[dict[str, Any]] = []
    selected_ids: set[str] = set()
    seen_styles: set[str] = set()

    for score, reasons, r in candidates:
        if len(selected) >= total:
            break
        if r["style"] in seen_styles:
            continue
        q = dict(r)
        q["selection_layer"] = "diagnostic_case_study"
        q["selection_reason"] = ";".join(reasons) or "diagnostic_case"
        q["diagnostic_case_score"] = round(score, 4)
        selected.append(q)
        selected_ids.add(r["doc_id"])
        seen_styles.add(r["style"])

    for score, reasons, r in candidates:
        if len(selected) >= total:
            break
        if r["doc_id"] in selected_ids:
            continue
        q = dict(r)
        q["selection_layer"] = "diagnostic_case_study"
        q["selection_reason"] = ";".join(reasons) or "diagnostic_case"
        q["diagnostic_case_score"] = round(score, 4)
        selected.append(q)
        selected_ids.add(r["doc_id"])

    return selected


def assign_reviewers(rows: list[dict[str, Any]], reviewers: int) -> list[dict[str, Any]]:
    ordered = sorted(rows, key=lambda r: (r.get("selection_layer", ""), r.get("style", ""), r.get("risk_band", ""), r.get("doc_id", "")))
    reviewer_counts = {i: 0 for i in range(1, reviewers + 1)}
    reviewer_style_counts: dict[int, dict[str, int]] = {i: defaultdict(int) for i in range(1, reviewers + 1)}

    assigned: list[dict[str, Any]] = []
    for r in ordered:
        style = r.get("style", "UNKNOWN")
        reviewer = min(reviewer_counts, key=lambda k: (reviewer_counts[k], reviewer_style_counts[k][style], k))
        reviewer_counts[reviewer] += 1
        reviewer_style_counts[reviewer][style] += 1
        q = dict(r)
        q["reviewer_id"] = f"reviewer_{reviewer}"
        q["review_order"] = reviewer_counts[reviewer]
        q["review_id"] = f"R{reviewer:02d}-{reviewer_counts[reviewer]:03d}"
        assigned.append(q)
    return assigned


# ---------------------------------------------------------------------
# Excel helpers
# ---------------------------------------------------------------------

def colour_symbol(hex_colour: str) -> str:
    return {
        FILL_GREEN: SYMBOL_GREEN,
        FILL_AMBER: SYMBOL_AMBER,
        FILL_RED: SYMBOL_RED,
        FILL_GREY: SYMBOL_GREY,
    }.get(hex_colour, "")


def option_text(entry: dict[str, str], lang: str = "es") -> str:
    symbol = colour_symbol(entry.get("colour", ""))
    prefix = f"{symbol} " if symbol else ""
    if lang == "en":
        return f"{prefix}{entry['label_en']} - {entry['meaning_en']}"
    return f"{prefix}{entry['label_es']} - {entry['meaning_es']}"


def excel_quote(text: str) -> str:
    return text.replace('"', '""')


def colour_name_es(hex_colour: str) -> str:
    return {
        FILL_GREEN: "Verde / baja preocupación",
        FILL_AMBER: "Ámbar / precaución",
        FILL_RED: "Rojo / alta preocupación",
        FILL_GREY: "Gris / pendiente o neutral",
    }.get(hex_colour, hex_colour)


def style_header(ws, max_col: int, title: str) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    c = ws.cell(1, 1)
    c.value = title
    c.fill = PatternFill("solid", fgColor=FILL_HEADER)
    c.font = Font(color="FFFFFF", bold=True, size=14)
    c.alignment = Alignment(horizontal="center")


def format_table_sheet(ws, header_row: int = 1, freeze_row: int = 2, widths: dict[int, int] | None = None) -> None:
    widths = widths or {}
    thin = Side(style="thin", color="D9E2F3")
    for cell in ws[header_row]:
        cell.fill = PatternFill("solid", fgColor=FILL_HEADER)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    ws.freeze_panes = f"A{freeze_row}"
    ws.auto_filter.ref = ws.dimensions
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = widths.get(col_idx, min(max(12, len(str(ws.cell(header_row, col_idx).value or "")) + 4), 48))
    for row in ws.iter_rows(min_row=header_row + 1):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)
    ws.sheet_view.showGridLines = False


def _safe_defined_name(category: str) -> str:
    return "opt_" + "".join(ch if ch.isalnum() or ch == "_" else "_" for ch in category)


def _add_defined_name(wb: Workbook, name: str, attr_text: str) -> None:
    try:
        if name in wb.defined_names:
            del wb.defined_names[name]
        wb.defined_names.add(DefinedName(name=name, attr_text=attr_text))
    except Exception:
        try:
            wb.defined_names.delete(name)
        except Exception:
            pass
        wb.defined_names.append(DefinedName(name=name, attr_text=attr_text))


def add_validation_lists_sheet(wb: Workbook) -> dict[str, str]:
    """Create hidden workbook-level validation lists for reviewer dropdowns.

    This restores the legacy validation mechanism: option lists live on a
    hidden worksheet and each review dimension is exposed through a workbook
    defined name. The visible review sheet then uses those defined names as
    data-validation sources. This avoids fragile inline-list limits and avoids
    direct same-sheet range formulas that can behave inconsistently across
    Excel/LibreOffice.
    """
    ws = wb.create_sheet("Listas_validacion")
    ranges: dict[str, str] = {}

    for col_idx, category in enumerate(REVIEW_DIMENSIONS, start=1):
        col_letter = get_column_letter(col_idx)
        ws.cell(1, col_idx).value = DIMENSION_SPANISH_HEADERS[category]
        ws.cell(1, col_idx).font = Font(bold=True)

        for row_idx, entry in enumerate(LABEL_SETS[category], start=2):
            ws.cell(row_idx, col_idx).value = option_text(entry, "es")
            ws.cell(row_idx, col_idx).fill = PatternFill("solid", fgColor=entry["colour"])

        last_row = len(LABEL_SETS[category]) + 1
        defined_name = _safe_defined_name(category)
        attr_text = f"'Listas_validacion'!${col_letter}$2:${col_letter}${last_row}"
        _add_defined_name(wb, defined_name, attr_text)
        ranges[category] = defined_name

        ws.column_dimensions[col_letter].width = 52

    ws.sheet_state = "hidden"
    return ranges

def add_label_definition_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Definiciones")
    headers = ["Categoría", "Definición / pregunta", "Etiqueta", "Descripción", "Texto del desplegable", "Semáforo"]
    ws.append(headers)
    for category in REVIEW_DIMENSIONS:
        for entry in LABEL_SETS[category]:
            ws.append([
                DIMENSION_SPANISH_HEADERS[category],
                DIMENSION_DESCRIPTIONS[category],
                entry["label_es"],
                entry["meaning_es"],
                option_text(entry, "es"),
                colour_name_es(entry["colour"]),
            ])
            row = ws.max_row
            ws.cell(row, 6).fill = PatternFill("solid", fgColor=entry["colour"])
    format_table_sheet(ws, header_row=1, freeze_row=2, widths={1: 36, 2: 95, 3: 42, 4: 76, 5: 110, 6: 24})


def add_domain_mapping_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Domain mappings")
    headers = [
        "domain_key", "Spanish domain", "English domain", "Spanish domain description",
        "English domain description", "Spanish label", "English label",
        "Spanish dropdown text", "English dropdown text",
    ]
    ws.append(headers)
    for category in REVIEW_DIMENSIONS:
        for entry in LABEL_SETS[category]:
            ws.append([
                category,
                DIMENSION_SPANISH_HEADERS[category],
                DIMENSION_ENGLISH_HEADERS[category],
                DIMENSION_DESCRIPTIONS[category],
                DIMENSION_DESCRIPTIONS_EN[category],
                entry["label_es"],
                entry["label_en"],
                option_text(entry, "es"),
                option_text(entry, "en"),
            ])
    format_table_sheet(ws, header_row=1, freeze_row=2, widths={1: 38, 2: 48, 3: 48, 4: 100, 5: 100, 6: 42, 7: 42, 8: 110, 9: 110})
    # This sheet is mainly for the project lead and is intentionally visible.


def apply_label_conditional_formatting(ws, col_letter: str, start_row: int, end_row: int, category: str) -> None:
    """Apply optional traffic-light formatting to reviewer dropdown cells.

    The dropdown values are prefixed with a traffic-light symbol, e.g.
    "🟢 Fiable - ...". The symbol is the primary cross-platform cue.
    Conditional formatting is kept as a nice-to-have for spreadsheet clients
    that support it.
    """
    target_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
    for entry in LABEL_SETS[category]:
        prefix_text = f"{colour_symbol(entry.get('colour', ''))} {entry['label_es']}".strip()
        prefix = excel_quote(prefix_text)
        prefix_len = len(prefix_text)
        formula = [f'LEFT(${col_letter}{start_row},{prefix_len})="{prefix}"']
        ws.conditional_formatting.add(
            target_range,
            FormulaRule(formula=formula, fill=PatternFill("solid", fgColor=entry["colour"])),
        )



def add_review_sheet(wb: Workbook, rows: list[dict[str, Any]], reviewer_id: str, validation_ranges: dict[str, str]) -> None:
    ws = wb.active
    ws.title = "Revisión"
    style_header(ws, len(SPANISH_COLUMNS), f"Paquete de revisión - {reviewer_id}")

    # Row 2: brief instruction for reviewers.
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(SPANISH_COLUMNS))
    ws.cell(2, 1).value = (
        "Seleccione una opción en cada columna de revisión. "
        "Los símbolos 🟢 🟡 🔴 ⚪ indican el nivel de preocupación/estado. "
        "Las columnas A–C son de solo lectura y no deben modificarse."
    )
    ws.cell(2, 1).fill = PatternFill("solid", fgColor=FILL_DESCRIPTION)
    ws.cell(2, 1).alignment = Alignment(wrap_text=True, vertical="top")

    # Row 3: Spanish headers.
    ws.append(SPANISH_COLUMNS)

    # Row 4: definitions/questions below headers. Keep this row tall enough
    # that the full definitions are visible on opening.
    desc_row = ["", "", ""]
    for category in REVIEW_DIMENSIONS:
        desc_row.append(DIMENSION_DESCRIPTIONS[category])
    ws.append(desc_row)

    for r in sorted(rows, key=lambda x: int(x.get("review_order", 0))):
        ws.append([
            r.get("review_id", ""),
            r.get("doc_id", ""),
            r.get("style", ""),
            "", "", "", "", "", "", "",
        ])

    # Formatting.
    widths = {
        1: 16, 2: 64, 3: 20, 4: 62, 5: 62, 6: 62,
        7: 62, 8: 62, 9: 48, 10: 48,
    }
    thin = Side(style="thin", color="D9E2F3")

    for cell in ws[3]:
        cell.fill = PatternFill("solid", fgColor=FILL_HEADER)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)

    for cell in ws[4]:
        cell.fill = PatternFill("solid", fgColor=FILL_DESCRIPTION)
        cell.font = Font(color="1F2328", italic=True, size=9)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = Border(bottom=thin)

    ws.freeze_panes = "D5"
    ws.auto_filter.ref = f"A3:{get_column_letter(ws.max_column)}{ws.max_row}"
    for col_idx in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(col_idx, 32)
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 44
    ws.row_dimensions[3].height = 44
    ws.row_dimensions[4].height = 75

    for row in ws.iter_rows(min_row=5):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=thin)

    # Visually mark identifier/context columns as read-only without using
    # worksheet protection, which behaves inconsistently in Excel Online.
    readonly_fill = PatternFill("solid", fgColor=FILL_READONLY)
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=3):
        for cell in row:
            cell.fill = readonly_fill

    max_row = max(ws.max_row, 5)

    # Data validation for review-domain cells. These start at row 5, below
    # the description row.
    for col_idx, category in zip(range(4, 11), REVIEW_DIMENSIONS):
        col_letter = get_column_letter(col_idx)
        dv = DataValidation(
            type="list",
            formula1=validation_ranges[category],
            allow_blank=True,
            showErrorMessage=True,
            errorTitle="Opción no válida",
            error="Seleccione una de las opciones del desplegable para esta categoría.",
        )
        # Leave showDropDown unset. In the OOXML spec this attribute has
        # inverted semantics and some spreadsheet clients misread
        # showDropDown="0". Omitting it gives the broadest compatibility.
        dv.showDropDown = None
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}5:{col_letter}{max_row}")
        apply_label_conditional_formatting(ws, col_letter, 5, max_row, category)



def make_reviewer_workbook(path: Path, rows: list[dict[str, Any]], reviewer_id: str) -> None:
    wb = Workbook()
    validation_ranges = add_validation_lists_sheet(wb)
    add_review_sheet(wb, rows, reviewer_id, validation_ranges)
    add_label_definition_sheet(wb)
    add_domain_mapping_sheet(wb)
    wb._sheets = [wb["Revisión"], wb["Definiciones"], wb["Domain mappings"], wb["Listas_validacion"]]
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


# ---------------------------------------------------------------------
# File-copy helpers for reviewer workspaces
# ---------------------------------------------------------------------

def safe_folder_name(value: str) -> str:
    value = str(value or "").strip()
    value = value.replace(":", "_")
    value = re.sub(r"[\\/*?\"<>|]", "_", value)
    value = re.sub(r"\s+", "_", value)
    return value[:180] or "documento_sin_id"


def resolve_source_path(row: dict[str, Any], kind: str) -> Path | None:
    """Resolve HTR/GT source path.

    Priority:
      1. explicit htr_path / gt_path from input CSV;
      2. explicit htr_path_or_url / gt_path_or_url;
      3. fallback reconstruction from expected filenames.
    """
    if kind == "htr":
        path_value = _first_non_empty(row, "htr_path", "htr_path_or_url", "HTR path", "HTR")
        expected_filename = row.get("expected_htr_filename", "")
    else:
        path_value = _first_non_empty(row, "gt_path", "gt_path_or_url", "GT path", "GT")
        expected_filename = row.get("expected_gt_filename", "")

    if path_value:
        p = Path(path_value)
        if p.exists() and p.is_file():
            return p

    # Fallback for GT if explicit path was not present.
    if kind == "gt" and expected_filename:
        # GT files are usually under Corpus_GT/<collection>/<file>.
        for p in GROUND_TRUTH_ROOT.rglob(expected_filename):
            if p.is_file():
                return p

    # Fallback for HTR: search raw data tree for expected HTR filename.
    if kind == "htr" and expected_filename:
        for p in RAW_DATA_ROOT.rglob(expected_filename):
            if p.is_file():
                return p

    return None


def copy_review_files_for_reviewer(packet_dir: Path, reviewer_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    warnings: list[dict[str, Any]] = []
    for r in sorted(reviewer_rows, key=lambda x: int(x.get("review_order", 0))):
        folder_name = safe_folder_name(str(r.get("packet_doc_id") or parse_compound_doc_id(r.get("doc_id", ""))[0]))
        doc_dir = packet_dir / folder_name
        doc_dir.mkdir(parents=True, exist_ok=True)

        htr_src = resolve_source_path(r, "htr")
        gt_src = resolve_source_path(r, "gt")

        if htr_src:
            shutil.copy2(htr_src, doc_dir / htr_src.name)
        else:
            warnings.append({
                "reviewer_id": r.get("reviewer_id", ""),
                "doc_id": r.get("doc_id", ""),
                "packet_doc_id": r.get("packet_doc_id", ""),
                "missing": "HTR",
                "htr_path": r.get("htr_path", r.get("htr_path_or_url", "")),
                "gt_path": r.get("gt_path", r.get("gt_path_or_url", "")),
                "expected_htr_filename": r.get("expected_htr_filename", ""),
                "expected_gt_filename": r.get("expected_gt_filename", ""),
            })

        if gt_src:
            shutil.copy2(gt_src, doc_dir / gt_src.name)
        else:
            warnings.append({
                "reviewer_id": r.get("reviewer_id", ""),
                "doc_id": r.get("doc_id", ""),
                "packet_doc_id": r.get("packet_doc_id", ""),
                "missing": "GT",
                "htr_path": r.get("htr_path", r.get("htr_path_or_url", "")),
                "gt_path": r.get("gt_path", r.get("gt_path_or_url", "")),
                "expected_htr_filename": r.get("expected_htr_filename", ""),
                "expected_gt_filename": r.get("expected_gt_filename", ""),
            })
    return warnings


# ---------------------------------------------------------------------
# Summary outputs
# ---------------------------------------------------------------------

def sample_summary(rows: list[dict[str, Any]]) -> dict[str, Any]:
    by_layer: dict[str, int] = defaultdict(int)
    by_style: dict[str, int] = defaultdict(int)
    by_reviewer: dict[str, int] = defaultdict(int)
    by_status: dict[str, int] = defaultdict(int)
    by_risk: dict[str, int] = defaultdict(int)
    for r in rows:
        by_layer[r.get("selection_layer", "")] += 1
        by_style[r.get("style", "")] += 1
        by_reviewer[r.get("reviewer_id", "")] += 1
        by_status[r.get("governance_status", r.get("document_status", ""))] += 1
        by_risk[r.get("risk_band", "")] += 1
    return {
        "total_documents": len(rows),
        "by_layer": dict(sorted(by_layer.items())),
        "by_style": dict(sorted(by_style.items())),
        "by_document_status": dict(sorted(by_status.items())),
        "by_risk_band": dict(sorted(by_risk.items())),
        "by_reviewer": dict(sorted(by_reviewer.items())),
    }


def html_escape(s: Any) -> str:
    return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def write_summary_html(path: Path, summary: dict[str, Any]) -> None:
    def table_from_dict(title: str, d: dict[str, Any]) -> str:
        trs = "".join(f"<tr><td>{html_escape(k)}</td><td>{v}</td></tr>" for k, v in d.items())
        return f"<h2>{html_escape(title)}</h2><table>{trs}</table>"

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>Human Review Sample Summary</title>
<style>
body{{font-family:system-ui,-apple-system,Segoe UI,Roboto;max-width:980px;margin:36px auto;line-height:1.55;color:#1f2328;background:#f7f7f8;}}
.card{{background:white;border:1px solid #d9d9de;border-radius:12px;padding:18px;margin:18px 0;}}
table{{border-collapse:collapse;width:100%;background:white;}}
th,td{{border:1px solid #e4e6eb;padding:7px 9px;text-align:left;}}
th{{background:#eef1f5;}}
</style>
</head>
<body>
<h1>Human Review Sample Summary</h1>
<div class="card">
<p><strong>Total documents:</strong> {summary['total_documents']}</p>
<p>Full document-level assignments are in <code>review_sample_all.csv</code>. Reviewer workspaces are in <code>reviewer_packets/</code>.</p>
</div>
<div class="card">{table_from_dict('Documents by layer', summary['by_layer'])}</div>
<div class="card">{table_from_dict('Documents by style', summary['by_style'])}</div>
<div class="card">{table_from_dict('Documents by document status', summary.get('by_document_status', {}))}</div>
<div class="card">{table_from_dict('Documents by risk band', summary.get('by_risk_band', {}))}</div>
<div class="card">{table_from_dict('Documents by reviewer', summary['by_reviewer'])}</div>
</body>
</html>"""
    path.write_text(html, encoding="utf-8")


# ---------------------------------------------------------------------
# Entrypoint
# ---------------------------------------------------------------------

def find_default_metrics_csv() -> Path:
    for candidate in DEFAULT_METRICS_CANDIDATES:
        if candidate.exists():
            return candidate
    raise SystemExit(
        "Could not find a default metrics CSV. Provide --metrics-csv. Tried: "
        + ", ".join(str(p) for p in DEFAULT_METRICS_CANDIDATES)
    )


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Build human-review samples and Spanish reviewer packets.")
    p.add_argument("--metrics-csv", type=Path, default=METRICS_CSV, help="Path to document_status_index.csv or operational metrics CSV")
    p.add_argument("--output-dir", type=Path, default=OUTPUT_DIR, help="Output directory")
    p.add_argument("--reviewers", type=int, default=NUMBER_OF_REVIEWERS, help="Number of reviewer packets to create")
    p.add_argument("--governance-total", type=int, default=GOVERNANCE_VALIDATION_TOTAL, help="Layer 1 target documents")
    p.add_argument("--controls-per-style", type=int, default=RANDOM_CONTROL_DOCS_PER_STYLE, help="Layer 2 random controls per style")
    p.add_argument("--diagnostic-total", type=int, default=DIAGNOSTIC_CASE_STUDY_TOTAL, help="Layer 3 diagnostic case studies")
    p.add_argument("--seed", type=int, default=RANDOM_SEED, help="Random seed for reproducibility")
    p.add_argument("--governance-method", choices=["even", "random", "extreme"], default=GOVERNANCE_SELECTION_METHOD)
    return p.parse_args()


def run_human_review_sampling() -> None:
    args = parse_args()
    rng = random.Random(args.seed)
    metrics_csv = args.metrics_csv or find_default_metrics_csv()
    output_dir: Path = args.output_dir
    packet_root = output_dir / "reviewer_packets"
    output_dir.mkdir(parents=True, exist_ok=True)
    packet_root.mkdir(parents=True, exist_ok=True)

    raw_rows = read_csv_dicts(metrics_csv)
    rows = normalise_doc_rows(raw_rows)
    if not rows:
        raise SystemExit(f"No usable document rows found in {metrics_csv}")

    usable_status_rows = [
        r for r in rows
        if r.get("governance_status") != "Unassigned" and r.get("risk_band") != "Unassigned"
    ]
    if not usable_status_rows:
        available_cols = sorted(raw_rows[0].keys()) if raw_rows else []
        raise SystemExit(
            "No usable style × document_status × risk_band cells were found. "
            "Expected fields such as doc_id, style, document_status/governance_status, and risk_band. "
            "Available columns: " + ", ".join(available_cols)
        )

    governance_rows = governance_validation_sample(rows, args.governance_total, args.governance_method, rng)
    used = {r["doc_id"] for r in governance_rows}

    control_rows = random_control_sample(rows, args.controls_per_style, used, rng)
    used.update(r["doc_id"] for r in control_rows)

    diagnostic_rows = diagnostic_case_studies(rows, args.diagnostic_total, used)

    all_rows = governance_rows + control_rows + diagnostic_rows
    assigned = assign_reviewers(all_rows, args.reviewers)

    write_csv_dicts(output_dir / "review_sample_all.csv", assigned, MASTER_COLUMNS)
    write_csv_dicts(output_dir / "review_sample_governance_validation.csv", [r for r in assigned if r["selection_layer"] == "governance_validation"], MASTER_COLUMNS)
    write_csv_dicts(output_dir / "review_sample_random_control.csv", [r for r in assigned if r["selection_layer"] == "random_control"], MASTER_COLUMNS)
    write_csv_dicts(output_dir / "review_sample_diagnostic_case_studies.csv", [r for r in assigned if r["selection_layer"] == "diagnostic_case_study"], MASTER_COLUMNS)

    by_reviewer: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for r in assigned:
        by_reviewer[r["reviewer_id"]].append(r)

    copy_warnings: list[dict[str, Any]] = []
    for reviewer_id, packet_rows in sorted(by_reviewer.items()):
        reviewer_packet_dir = packet_root / f"{reviewer_id}_packet"
        if reviewer_packet_dir.exists():
            shutil.rmtree(reviewer_packet_dir)
        reviewer_packet_dir.mkdir(parents=True, exist_ok=True)

        make_reviewer_workbook(reviewer_packet_dir / f"{reviewer_id}_packet.xlsx", packet_rows, reviewer_id)

        # CSV companion mirrors the Spanish review sheet using internal column names
        # for easier later import.
        csv_rows = []
        for r in sorted(packet_rows, key=lambda x: int(x.get("review_order", 0))):
            row = {
                "review_id": r.get("review_id", ""),
                "doc_id": r.get("doc_id", ""),
                "style": r.get("style", ""),
                **{dim: "" for dim in REVIEW_DIMENSIONS},
            }
            csv_rows.append(row)
        write_csv_dicts(reviewer_packet_dir / f"{reviewer_id}_packet.csv", csv_rows, REVIEW_COLUMNS_INTERNAL)

        copy_warnings.extend(copy_review_files_for_reviewer(reviewer_packet_dir, packet_rows))

    warning_cols = [
        "reviewer_id", "doc_id", "packet_doc_id", "missing",
        "htr_path", "gt_path", "expected_htr_filename", "expected_gt_filename",
    ]
    write_csv_dicts(output_dir / "file_copy_warnings.csv", copy_warnings, warning_cols)

    label_schema = {
        "label_sets": {
            k: [{**entry, "dropdown_text_es": option_text(entry, "es"), "dropdown_text_en": option_text(entry, "en")} for entry in v]
            for k, v in LABEL_SETS.items()
        },
        "review_columns_internal": REVIEW_COLUMNS_INTERNAL,
        "review_columns_spanish": SPANISH_COLUMNS,
        "symbol_legend": {
            "🟢": "baja preocupación / opción positiva / completado",
            "🟡": "precaución / estado intermedio",
            "🔴": "alta preocupación / problema significativo",
            "⚪": "pendiente o neutral",
        },
        "sampling_design": {
            "governance_validation": (
                f"target {args.governance_total} documents selected round-robin from "
                f"style × governance_status × risk_band cells, method={args.governance_method}"
            ),
            "random_control": f"{args.controls_per_style} per style, excluding governance-validation docs",
            "diagnostic_case_studies": f"{args.diagnostic_total} purposive diagnostic case studies, excluding prior layers",
            "seed": args.seed,
        },
    }
    (output_dir / "review_label_schema.json").write_text(json.dumps(label_schema, ensure_ascii=False, indent=2), encoding="utf-8")

    summary = sample_summary(assigned)
    summary["file_copy_warnings"] = len(copy_warnings)
    (output_dir / "review_sample_summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    write_summary_html(output_dir / "review_sample_summary.html", summary)

    print(f"Metrics source: {metrics_csv}")
    print(f"Output directory: {output_dir}")
    print(json.dumps(summary, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    run_human_review_sampling()
